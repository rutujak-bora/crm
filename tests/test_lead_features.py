"""
Test Lead Features: Part Number and Tender Document
Tests for the new Part Number (product level) and Tender Document (lead level) features
"""
import pytest
import requests
import os
import tempfile
from datetime import datetime

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
TEST_EMAIL = "sunil@bora.tech"
TEST_PASSWORD = "sunil@1202"


class TestAuth:
    """Authentication tests"""
    
    def test_login_success(self):
        """Test successful login"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": TEST_EMAIL,
            "password": TEST_PASSWORD
        })
        assert response.status_code == 200, f"Login failed: {response.text}"
        data = response.json()
        assert "token" in data
        assert "user" in data
        assert data["user"]["email"] == TEST_EMAIL
        print(f"✓ Login successful, token received")
        return data["token"]


@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token for tests"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": TEST_EMAIL,
        "password": TEST_PASSWORD
    })
    if response.status_code == 200:
        return response.json()["token"]
    pytest.skip("Authentication failed")


@pytest.fixture(scope="module")
def auth_headers(auth_token):
    """Get auth headers"""
    return {"Authorization": f"Bearer {auth_token}"}


@pytest.fixture(scope="module")
def test_customer(auth_headers):
    """Create a test customer for lead tests"""
    customer_data = {
        "customer_name": f"TEST_Customer_{datetime.now().strftime('%H%M%S')}",
        "reference_name": "TEST_REF",
        "contact_number": "9876543210",
        "email": "test@example.com"
    }
    response = requests.post(f"{BASE_URL}/api/customers", json=customer_data, headers=auth_headers)
    assert response.status_code == 200, f"Failed to create customer: {response.text}"
    customer = response.json()
    print(f"✓ Created test customer: {customer['customer_name']}")
    yield customer
    # Cleanup
    requests.delete(f"{BASE_URL}/api/customers/{customer['id']}", headers=auth_headers)
    print(f"✓ Cleaned up test customer")


class TestPartNumber:
    """Tests for Part Number field at product level"""
    
    def test_create_lead_with_part_number(self, auth_headers, test_customer):
        """Test creating a lead with part_number in products"""
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [
                {
                    "product": "TEST_Product_A",
                    "part_number": "PN-001-TEST",
                    "category": "Electronics",
                    "quantity": 10,
                    "price": 100
                },
                {
                    "product": "TEST_Product_B",
                    "part_number": "PN-002-TEST",
                    "category": "Hardware",
                    "quantity": 5,
                    "price": 200
                }
            ]
        }
        
        response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        assert response.status_code == 200, f"Failed to create lead: {response.text}"
        
        lead = response.json()
        assert "id" in lead
        assert len(lead["products"]) == 2
        
        # Verify part_number is saved correctly
        assert lead["products"][0]["part_number"] == "PN-001-TEST"
        assert lead["products"][1]["part_number"] == "PN-002-TEST"
        
        # Verify amounts are calculated
        assert lead["products"][0]["amount"] == 1000  # 10 * 100
        assert lead["products"][1]["amount"] == 1000  # 5 * 200
        assert lead["total_amount"] == 2000
        
        print(f"✓ Lead created with part numbers: {lead['id']}")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/leads/{lead['id']}", headers=auth_headers)
        print(f"✓ Cleaned up test lead")
    
    def test_create_lead_without_part_number(self, auth_headers, test_customer):
        """Test creating a lead without part_number (should be optional)"""
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [
                {
                    "product": "TEST_Product_No_PN",
                    "category": "General",
                    "quantity": 3,
                    "price": 50
                }
            ]
        }
        
        response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        assert response.status_code == 200, f"Failed to create lead: {response.text}"
        
        lead = response.json()
        # part_number should be None or not present
        assert lead["products"][0].get("part_number") is None or lead["products"][0].get("part_number") == ""
        print(f"✓ Lead created without part number (optional field works)")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/leads/{lead['id']}", headers=auth_headers)
    
    def test_update_lead_part_number(self, auth_headers, test_customer):
        """Test updating part_number in an existing lead"""
        # Create lead first
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [
                {
                    "product": "TEST_Product_Update",
                    "part_number": "PN-ORIGINAL",
                    "category": "Test",
                    "quantity": 1,
                    "price": 100
                }
            ]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        assert create_response.status_code == 200
        lead = create_response.json()
        lead_id = lead["id"]
        
        # Update the lead with new part_number
        update_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [
                {
                    "product": "TEST_Product_Update",
                    "part_number": "PN-UPDATED",
                    "category": "Test",
                    "quantity": 2,
                    "price": 150
                }
            ]
        }
        
        update_response = requests.put(f"{BASE_URL}/api/leads/{lead_id}", json=update_data, headers=auth_headers)
        assert update_response.status_code == 200, f"Failed to update lead: {update_response.text}"
        
        updated_lead = update_response.json()
        assert updated_lead["products"][0]["part_number"] == "PN-UPDATED"
        assert updated_lead["products"][0]["quantity"] == 2
        assert updated_lead["products"][0]["price"] == 150
        
        print(f"✓ Lead part_number updated successfully")
        
        # Verify with GET
        get_response = requests.get(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
        assert get_response.status_code == 200
        fetched_lead = get_response.json()
        assert fetched_lead["products"][0]["part_number"] == "PN-UPDATED"
        print(f"✓ Part number persisted correctly after update")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
    
    def test_get_lead_with_part_number(self, auth_headers, test_customer):
        """Test retrieving a lead shows part_number correctly"""
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [
                {
                    "product": "TEST_Product_Get",
                    "part_number": "PN-GET-TEST",
                    "category": "Test",
                    "quantity": 5,
                    "price": 75
                }
            ]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        lead = create_response.json()
        lead_id = lead["id"]
        
        # GET the lead
        get_response = requests.get(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
        assert get_response.status_code == 200
        
        fetched_lead = get_response.json()
        assert fetched_lead["products"][0]["part_number"] == "PN-GET-TEST"
        print(f"✓ GET lead returns part_number correctly")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)


class TestTenderDocument:
    """Tests for Tender Document upload/download at lead level"""
    
    def test_upload_tender_document_pdf(self, auth_headers, test_customer):
        """Test uploading a PDF tender document"""
        # Create a lead first
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [{"product": "TEST_Product", "category": "Test", "quantity": 1, "price": 100}]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        lead = create_response.json()
        lead_id = lead["id"]
        
        # Create a test PDF file
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(b"%PDF-1.4 test content")
            temp_file_path = f.name
        
        try:
            with open(temp_file_path, "rb") as f:
                files = {"file": ("test_tender.pdf", f, "application/pdf")}
                upload_response = requests.post(
                    f"{BASE_URL}/api/leads/{lead_id}/upload-tender-document",
                    files=files,
                    headers=auth_headers
                )
            
            assert upload_response.status_code == 200, f"Upload failed: {upload_response.text}"
            upload_data = upload_response.json()
            assert "document_url" in upload_data
            assert upload_data["document_url"].startswith("/api/uploads/")
            print(f"✓ PDF tender document uploaded: {upload_data['document_url']}")
            
            # Verify lead has tender_document
            get_response = requests.get(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
            lead_data = get_response.json()
            assert lead_data["tender_document"] == upload_data["document_url"]
            print(f"✓ Lead tender_document field updated correctly")
            
            # Test downloading the document
            download_response = requests.get(f"{BASE_URL}{upload_data['document_url']}")
            assert download_response.status_code == 200
            print(f"✓ Tender document can be downloaded")
            
        finally:
            os.unlink(temp_file_path)
            requests.delete(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
    
    def test_upload_tender_document_docx(self, auth_headers, test_customer):
        """Test uploading a DOCX tender document"""
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [{"product": "TEST_Product", "category": "Test", "quantity": 1, "price": 100}]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        lead = create_response.json()
        lead_id = lead["id"]
        
        # Create a test DOCX file (minimal valid docx is complex, so we just test the extension)
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as f:
            f.write(b"PK test docx content")
            temp_file_path = f.name
        
        try:
            with open(temp_file_path, "rb") as f:
                files = {"file": ("test_tender.docx", f, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")}
                upload_response = requests.post(
                    f"{BASE_URL}/api/leads/{lead_id}/upload-tender-document",
                    files=files,
                    headers=auth_headers
                )
            
            assert upload_response.status_code == 200, f"DOCX upload failed: {upload_response.text}"
            print(f"✓ DOCX tender document uploaded successfully")
            
        finally:
            os.unlink(temp_file_path)
            requests.delete(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
    
    def test_upload_tender_document_xlsx(self, auth_headers, test_customer):
        """Test uploading an XLSX tender document"""
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [{"product": "TEST_Product", "category": "Test", "quantity": 1, "price": 100}]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        lead = create_response.json()
        lead_id = lead["id"]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"PK test xlsx content")
            temp_file_path = f.name
        
        try:
            with open(temp_file_path, "rb") as f:
                files = {"file": ("test_tender.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                upload_response = requests.post(
                    f"{BASE_URL}/api/leads/{lead_id}/upload-tender-document",
                    files=files,
                    headers=auth_headers
                )
            
            assert upload_response.status_code == 200, f"XLSX upload failed: {upload_response.text}"
            print(f"✓ XLSX tender document uploaded successfully")
            
        finally:
            os.unlink(temp_file_path)
            requests.delete(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
    
    def test_upload_tender_document_png(self, auth_headers, test_customer):
        """Test uploading a PNG tender document"""
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [{"product": "TEST_Product", "category": "Test", "quantity": 1, "price": 100}]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        lead = create_response.json()
        lead_id = lead["id"]
        
        # Create a minimal valid PNG
        png_header = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde'
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
            f.write(png_header)
            temp_file_path = f.name
        
        try:
            with open(temp_file_path, "rb") as f:
                files = {"file": ("test_tender.png", f, "image/png")}
                upload_response = requests.post(
                    f"{BASE_URL}/api/leads/{lead_id}/upload-tender-document",
                    files=files,
                    headers=auth_headers
                )
            
            assert upload_response.status_code == 200, f"PNG upload failed: {upload_response.text}"
            print(f"✓ PNG tender document uploaded successfully")
            
        finally:
            os.unlink(temp_file_path)
            requests.delete(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
    
    def test_upload_invalid_file_type(self, auth_headers, test_customer):
        """Test that invalid file types are rejected"""
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [{"product": "TEST_Product", "category": "Test", "quantity": 1, "price": 100}]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        lead = create_response.json()
        lead_id = lead["id"]
        
        with tempfile.NamedTemporaryFile(suffix=".exe", delete=False) as f:
            f.write(b"invalid file content")
            temp_file_path = f.name
        
        try:
            with open(temp_file_path, "rb") as f:
                files = {"file": ("test.exe", f, "application/octet-stream")}
                upload_response = requests.post(
                    f"{BASE_URL}/api/leads/{lead_id}/upload-tender-document",
                    files=files,
                    headers=auth_headers
                )
            
            assert upload_response.status_code == 400, f"Expected 400 for invalid file type, got {upload_response.status_code}"
            print(f"✓ Invalid file type (.exe) correctly rejected")
            
        finally:
            os.unlink(temp_file_path)
            requests.delete(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
    
    def test_delete_tender_document(self, auth_headers, test_customer):
        """Test deleting a tender document from a lead"""
        # Create lead and upload document
        lead_data = {
            "customer_id": test_customer["id"],
            "customer_name": test_customer["customer_name"],
            "date": datetime.now().strftime("%Y-%m-%d"),
            "products": [{"product": "TEST_Product", "category": "Test", "quantity": 1, "price": 100}]
        }
        
        create_response = requests.post(f"{BASE_URL}/api/leads", json=lead_data, headers=auth_headers)
        lead = create_response.json()
        lead_id = lead["id"]
        
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(b"%PDF-1.4 test content")
            temp_file_path = f.name
        
        try:
            with open(temp_file_path, "rb") as f:
                files = {"file": ("test_tender.pdf", f, "application/pdf")}
                requests.post(
                    f"{BASE_URL}/api/leads/{lead_id}/upload-tender-document",
                    files=files,
                    headers=auth_headers
                )
            
            # Delete the tender document
            delete_response = requests.delete(
                f"{BASE_URL}/api/leads/{lead_id}/tender-document",
                headers=auth_headers
            )
            assert delete_response.status_code == 200, f"Delete failed: {delete_response.text}"
            print(f"✓ Tender document deleted successfully")
            
            # Verify lead no longer has tender_document
            get_response = requests.get(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
            lead_data = get_response.json()
            assert lead_data["tender_document"] is None
            print(f"✓ Lead tender_document field cleared after delete")
            
        finally:
            os.unlink(temp_file_path)
            requests.delete(f"{BASE_URL}/api/leads/{lead_id}", headers=auth_headers)
    
    def test_upload_to_nonexistent_lead(self, auth_headers):
        """Test uploading to a non-existent lead returns 404"""
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(b"%PDF-1.4 test content")
            temp_file_path = f.name
        
        try:
            with open(temp_file_path, "rb") as f:
                files = {"file": ("test.pdf", f, "application/pdf")}
                upload_response = requests.post(
                    f"{BASE_URL}/api/leads/nonexistent-id/upload-tender-document",
                    files=files,
                    headers=auth_headers
                )
            
            assert upload_response.status_code == 404
            print(f"✓ Upload to non-existent lead correctly returns 404")
            
        finally:
            os.unlink(temp_file_path)


class TestExcelTemplate:
    """Tests for Excel template with Part Number column"""
    
    def test_download_lead_template(self, auth_headers):
        """Test downloading lead template includes Part Number column"""
        response = requests.get(f"{BASE_URL}/api/leads/template/download", headers=auth_headers)
        assert response.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("content-type", "")
        
        # Save and verify content
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(response.content)
            temp_file_path = f.name
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_file_path)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Verify Part Number column exists
            assert "Part Number" in headers, f"Part Number column not found in headers: {headers}"
            print(f"✓ Lead template contains Part Number column")
            print(f"  Headers: {headers}")
            
            # Verify column order: Part Number should be after Product
            product_idx = headers.index("Product*")
            part_number_idx = headers.index("Part Number")
            assert part_number_idx == product_idx + 1, "Part Number should be right after Product"
            print(f"✓ Part Number column is correctly positioned after Product")
            
        finally:
            os.unlink(temp_file_path)


class TestBulkUpload:
    """Tests for bulk upload with Part Number"""
    
    def test_bulk_upload_with_part_number(self, auth_headers, test_customer):
        """Test bulk upload processes Part Number correctly"""
        from openpyxl import Workbook
        
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.append(["Customer Name*", "Proforma Invoice No", "Date*", "Product*", "Part Number", "Category*", "Quantity*", "Price*", "Follow-up Date", "Remark"])
        ws.append([test_customer["customer_name"], "TEST-PI-001", datetime.now().strftime("%Y-%m-%d"), "Bulk Product A", "BULK-PN-001", "Electronics", 10, 100, "", ""])
        ws.append([test_customer["customer_name"], "TEST-PI-001", datetime.now().strftime("%Y-%m-%d"), "Bulk Product B", "BULK-PN-002", "Hardware", 5, 200, "", ""])
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_file_path = f.name
        
        try:
            with open(temp_file_path, "rb") as f:
                files = {"file": ("leads.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                upload_response = requests.post(
                    f"{BASE_URL}/api/leads/bulk-upload",
                    files=files,
                    headers=auth_headers
                )
            
            assert upload_response.status_code == 200, f"Bulk upload failed: {upload_response.text}"
            result = upload_response.json()
            assert result["created"] >= 1, f"Expected at least 1 lead created, got {result['created']}"
            print(f"✓ Bulk upload created {result['created']} lead(s)")
            
            # Verify the lead has part numbers
            leads_response = requests.get(f"{BASE_URL}/api/leads", headers=auth_headers)
            leads = leads_response.json()
            
            # Find the lead we just created
            test_lead = None
            for lead in leads:
                if lead.get("proforma_invoice_number") == "TEST-PI-001":
                    test_lead = lead
                    break
            
            if test_lead:
                # Verify part numbers
                part_numbers = [p.get("part_number") for p in test_lead["products"]]
                assert "BULK-PN-001" in part_numbers, f"Part number BULK-PN-001 not found in {part_numbers}"
                assert "BULK-PN-002" in part_numbers, f"Part number BULK-PN-002 not found in {part_numbers}"
                print(f"✓ Bulk uploaded lead contains correct part numbers: {part_numbers}")
                
                # Cleanup
                requests.delete(f"{BASE_URL}/api/leads/{test_lead['id']}", headers=auth_headers)
            else:
                print("⚠ Could not find test lead to verify part numbers")
            
        finally:
            os.unlink(temp_file_path)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
