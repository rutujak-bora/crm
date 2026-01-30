from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import io
from openpyxl import Workbook, load_workbook
import shutil
from dotenv import load_dotenv
import urllib.parse
import asyncio
import certifi
from bid_reminder_scheduler import init_scheduler, shutdown_scheduler, get_scheduler_status

# ================= SETUP & CONFIG =================

# 1. Environment & Directories
BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / ".env"
load_dotenv(ENV_PATH)

ROOT_DIR = Path(__file__).parent
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
GEM_BID_UPLOAD_DIR = ROOT_DIR / "gem_uploads"
GEM_BID_UPLOAD_DIR.mkdir(exist_ok=True)

# 1.5 Default Credentials
CRM_USER_EMAIL = os.getenv("CRM_USER_EMAIL", "sunil@bora.tech")
CRM_USER_PASSWORD = os.getenv("CRM_USER_PASSWORD", "sunil@1202")
GEM_BID_USER_EMAIL = os.getenv("GEM_BID_USER_EMAIL", "yash.b@bora.tech")
GEM_BID_USER_PASSWORD = os.getenv("GEM_BID_USER_PASSWORD", "yash@123")

# 2. Database Configuration
raw_uri = os.getenv("MONGO_URI") or os.getenv("MONGO_URL")
DB_NAME = os.getenv("DB_NAME")

MONGO_URI = raw_uri
if raw_uri and "mongodb" in raw_uri and "@" in raw_uri:
    try:
        prefix = "mongodb+srv://" if "mongodb+srv://" in raw_uri else "mongodb://"
        creds_host = raw_uri.replace(prefix, "")
        if "@" in creds_host:
            creds, host = creds_host.rsplit("@", 1)
            if ":" in creds:
                user, pwd = creds.split(":", 1)
                pwd = urllib.parse.unquote(pwd)
                encoded_pwd = urllib.parse.quote_plus(pwd)
                MONGO_URI = f"{prefix}{user}:{encoded_pwd}@{host}"
    except Exception as e:
        print(f"Error parsing MONGO_URI: {e}")

# --- Mock DB for Demo Mode ---
class MockCollection:
    def __init__(self, name):
        self.name = name
        self.data = []
    async def find_one(self, query, projection=None):
        for item in self.data:
            if self._match(item, query):
                return item.copy()
        return None

    def _match(self, item, query):
        if not query: return True
        for k, v in query.items():
            if k == "$ne": # Handle basic $ne for MockDB
                continue 
            if isinstance(v, dict):
                if "$ne" in v:
                    if item.get(k) == v["$ne"]: return False
                else:
                    if item.get(k) != v: return False
            elif item.get(k) != v:
                return False
        return True

    def find(self, query=None, projection=None):
        class MockCursor:
            def __init__(self, data): self.data = data
            def sort(self, *args, **kwargs):
                # Basic mock sort - doesn't actually sort but matches API
                return self
            async def to_list(self, length): return self.data
        
        res = [item.copy() for item in self.data if self._match(item, query)]
        return MockCursor(res)
    
    async def insert_one(self, doc):
        self.data.append(doc)
        return type('obj', (), {'inserted_id': 'demo_id'})
    
    async def insert_many(self, docs):
        self.data.extend(docs)
        return True
    
    async def update_one(self, query, update, upsert=False):
        for item in self.data:
            if self._match(item, query):
                if "$set" in update: item.update(update["$set"])
                return True
        if upsert: await self.insert_one({**query, **update.get("$set", {})})
        return True
    
    async def delete_one(self, query):
        for i, item in enumerate(self.data):
            if self._match(item, query):
                self.data.pop(i)
                return type('obj', (), {'deleted_count': 1})
        return type('obj', (), {'deleted_count': 0})
    
    async def count_documents(self, query):
        return len([item for item in self.data if self._match(item, query)])

    async def aggregate(self, pipeline): 
        class MockCursor:
            async def to_list(self, length): return []
        return MockCursor()

class MockDB:
    def __init__(self):
        self.collections = {}
    def __getattr__(self, name):
        if name not in self.collections: self.collections[name] = MockCollection(name)
        return self.collections[name]
    def __getitem__(self, name): return getattr(self, name)
    async def command(self, cmd):
        if cmd == 'ping': return True
        return True

# Initialize DB with fallback
try:
    if not MONGO_URI or not DB_NAME:
        db = MockDB()
        logger.warning("No MONGO_URI. Using Demo Mode (In-memory)")
    else:
        mongo_client = AsyncIOMotorClient(MONGO_URI, serverSelectionTimeoutMS=5000, tlsCAFile=certifi.where())
        db = mongo_client[DB_NAME]
except Exception:
    db = MockDB()
    logger.warning("Database init failed. Using Demo Mode (In-memory)")

# 3. Security / Auth Configuration
SECRET_KEY = os.environ.get('JWT_SECRET', 'crm-secret-key-2024')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

security = HTTPBearer()

# 4. App & Logging Initialization
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = FastAPI(title="CRM Backend")
api_router = APIRouter(prefix="/api")

# Startup event to ensure users exist in database
@app.on_event("startup")
async def startup_event():
    global db
    try:
        # Check connection with a longer timeout
        await asyncio.wait_for(db.command('ping'), timeout=5.0)
        logger.info("MongoDB connected successfully")
    except Exception as e:
        logger.error(f"MongoDB connection failed: {e}. Switching to Demo Mode (In-Memory).")
        db = MockDB()
    
    try:
        # Check if users collection exists and has data
        count = await db.users.count_documents({})
        if count == 0:
            logger.info("Initializing users collection with default credentials")
            default_users = [
                {
                    "email": CRM_USER_EMAIL,
                    "password": CRM_USER_PASSWORD,
                    "name": "Sunil Bora",
                    "system": "crm"
                },
                {
                    "email": GEM_BID_USER_EMAIL,
                    "password": GEM_BID_USER_PASSWORD,
                    "name": "Yash Bora",
                    "system": "gem_bid"
                }
            ]
            await db.users.insert_many(default_users)
        else:
            # Sync passwords from environment to database
            await db.users.update_one(
                {"email": CRM_USER_EMAIL, "system": "crm"},
                {"$set": {"password": CRM_USER_PASSWORD}}
            )
            await db.users.update_one(
                {"email": GEM_BID_USER_EMAIL, "system": "gem_bid"},
                {"$set": {"password": GEM_BID_USER_PASSWORD}}
            )
        logger.info("Database initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize users in database: {e}")
    
    # Initialize bid reminder scheduler
    try:
        init_scheduler(db)
        logger.info("Bid reminder scheduler initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize bid reminder scheduler: {e}")

# Allowed file types for documents
ALLOWED_EXTENSIONS = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.png'}
MAX_FILE_SIZE = 25 * 1024 * 1024  # 25MB

@app.get("/")
def root():
    return {"message": "CRM Backend Running"}

@app.get("/test-db")
async def test_db():
    try:
        # Just a check command
        await db.command('ping')
        return {"status": "MongoDB Connected", "database": DB_NAME}
    except Exception as e:
        return {"status": "MongoDB Connection Failed", "error": str(e)}

# ============== MODELS ==============

class LoginRequest(BaseModel):
    email: str
    password: str

class LoginResponse(BaseModel):
    token: str
    user: dict

class CustomerBase(BaseModel):
    customer_name: str
    reference_name: Optional[str] = None
    contact_number: str
    email: str

class CustomerCreate(CustomerBase):
    pass

class Customer(CustomerBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ProductItem(BaseModel):
    product: str
    part_number: Optional[str] = None
    category: str
    quantity: float
    price: float
    amount: float = 0

class LeadBase(BaseModel):
    customer_id: str
    customer_name: str
    proforma_invoice_number: Optional[str] = None
    date: str
    products: List[ProductItem]
    follow_up_date: Optional[str] = None
    remark: Optional[str] = None
    tender_document: Optional[str] = None  # File path/URL for tender document (customer-provided)
    working_sheet: Optional[str] = None  # File path/URL for working sheet (internal reference)
    is_converted: bool = False

class LeadCreate(LeadBase):
    pass

class Lead(LeadBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    total_amount: float = 0

class ProformaInvoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    proforma_invoice_number: str
    customer_id: str
    customer_name: str
    date: str
    products: List[ProductItem]
    total_amount: float
    tender_document: Optional[str] = None  # Synced from lead
    working_sheet: Optional[str] = None  # Synced from lead
    created_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    lead_id: str

class POProductItem(BaseModel):
    product: str
    category: str
    quantity: float
    price: float
    amount: float = 0

class PurchaseOrderBase(BaseModel):
    purchase_order_number: str
    date: str
    vendor_name: str
    purpose: str  # "linked" or "stock_in_sale"
    proforma_invoice_id: Optional[str] = None
    proforma_invoice_number: Optional[str] = None
    products: List[POProductItem] = []

class PurchaseOrderCreate(PurchaseOrderBase):
    pass

class PurchaseOrder(PurchaseOrderBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    total_amount: float = 0
    created_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    # Backward compatibility fields
    product: Optional[str] = None
    category: Optional[str] = None
    quantity: Optional[float] = None
    price: Optional[float] = None
    amount: Optional[float] = None

class MarginEntry(BaseModel):
    proforma_invoice_number: str
    proforma_invoice_id: str
    proforma_total_amount: float
    purchase_order_number: str
    purchase_order_id: str
    purchase_order_amount: float
    remaining_amount: float
    freight_amount: float = 0
    margin_amount: float = 0

class MarginUpdate(BaseModel):
    freight_amount: float

# ============== AUTH ==============

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("email")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token")
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

@api_router.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    # Hardcoded fallback from environment
    HARDCODED_CRM = {"email": CRM_USER_EMAIL, "password": CRM_USER_PASSWORD, "name": "Sunil Bora"}
    
    try:
        user = await db.users.find_one({
            "email": request.email, 
            "password": request.password,
            "system": "crm"
        })
        if user:
            token = create_access_token({"email": user["email"], "name": user["name"]})
            return {"token": token, "user": {"email": user["email"], "name": user["name"]}}
    except Exception as e:
        logger.error(f"Database login error: {e}")
    
    # Check hardcoded fallback
    if request.email == HARDCODED_CRM["email"] and request.password == HARDCODED_CRM["password"]:
        token = create_access_token({"email": HARDCODED_CRM["email"], "name": HARDCODED_CRM["name"]})
        return {"token": token, "user": {"email": HARDCODED_CRM["email"], "name": HARDCODED_CRM["name"]}}
        
    raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.get("/auth/verify")
async def verify_auth(user: dict = Depends(verify_token)):
    return {"valid": True, "user": user}

# ============== DASHBOARD ==============

@api_router.get("/dashboard/kpi")
async def get_dashboard_kpi(user: dict = Depends(verify_token)):
    total_customers = await db.customers.count_documents({})
    active_leads = await db.leads.count_documents({"is_converted": False})
    total_proforma = await db.proforma_invoices.count_documents({})
    total_purchase_orders = await db.purchase_orders.count_documents({})
    
    # Calculate margin summary
    margin_pipeline = [
        {"$lookup": {
            "from": "purchase_orders",
            "localField": "id",
            "foreignField": "proforma_invoice_id",
            "as": "linked_orders"
        }},
        {"$match": {"linked_orders.0": {"$exists": True}}}
    ]
    proforma_with_orders = await db.proforma_invoices.aggregate(margin_pipeline).to_list(1000)
    
    total_margin = 0
    for pf in proforma_with_orders:
        po_total = sum(po.get("total_amount", po.get("amount", 0)) for po in pf.get("linked_orders", []))
        remaining = pf.get("total_amount", 0) - po_total
        # Margin is remaining minus freight (assume 0 freight for summary)
        total_margin += remaining
    
    return {
        "total_customers": total_customers,
        "active_leads": active_leads,
        "total_proforma_invoices": total_proforma,
        "total_purchase_orders": total_purchase_orders,
        "margin_summary": round(total_margin, 2)
    }

# ============== CUSTOMERS ==============

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer: CustomerCreate, user: dict = Depends(verify_token)):
    customer_obj = Customer(**customer.model_dump())
    doc = customer_obj.model_dump()
    await db.customers.insert_one(doc)
    return customer_obj

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(user: dict = Depends(verify_token)):
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    return customers

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, user: dict = Depends(verify_token)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer: CustomerCreate, user: dict = Depends(verify_token)):
    existing = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer not found")
    update_data = customer.model_dump()
    await db.customers.update_one({"id": customer_id}, {"$set": update_data})
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    return updated

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, user: dict = Depends(verify_token)):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted"}

@api_router.post("/customers/bulk-upload")
async def bulk_upload_customers(file: UploadFile = File(...), user: dict = Depends(verify_token)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    customers_created = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        try:
            customer = Customer(
                customer_name=str(row[0]),
                reference_name=str(row[1]) if row[1] else None,
                contact_number=str(row[2]) if row[2] else "",
                email=str(row[3]) if row[3] else ""
            )
            doc = customer.model_dump()
            await db.customers.insert_one(doc)
            customers_created += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return {"created": customers_created, "errors": errors}

@api_router.get("/customers/template/download")
async def download_customer_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = ["Customer Name*", "Reference Name", "Contact Number*", "Email*"]
    ws.append(headers)
    ws.append(["John Doe", "REF001", "9876543210", "john@example.com"])
    ws.append(["Jane Smith", "REF002", "9876543211", "jane@example.com"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_template.xlsx"}
    )

# ============== LEADS ==============

@api_router.post("/leads", response_model=Lead)
async def create_lead(lead: LeadCreate, user: dict = Depends(verify_token)):
    # Validate customer exists
    customer = await db.customers.find_one({"id": lead.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=400, detail="Customer not found")
    
    # Calculate amounts
    products = []
    total_amount = 0
    for p in lead.products:
        amount = p.quantity * p.price
        products.append(ProductItem(
            product=p.product,
            part_number=p.part_number,
            category=p.category,
            quantity=p.quantity,
            price=p.price,
            amount=round(amount, 2)
        ))
        total_amount += amount
    
    lead_obj = Lead(
        **lead.model_dump(exclude={"products"}),
        products=products,
        total_amount=round(total_amount, 2)
    )
    doc = lead_obj.model_dump()
    await db.leads.insert_one(doc)
    return lead_obj

@api_router.get("/leads", response_model=List[Lead])
async def get_leads(
    customer_name: Optional[str] = None,
    category: Optional[str] = None,
    user: dict = Depends(verify_token)
):
    query = {}
    if customer_name:
        query["customer_name"] = {"$regex": customer_name, "$options": "i"}
    if category:
        query["products.category"] = {"$regex": category, "$options": "i"}
    
    leads = await db.leads.find(query, {"_id": 0}).to_list(1000)
    return leads

@api_router.get("/leads/{lead_id}", response_model=Lead)
async def get_lead(lead_id: str, user: dict = Depends(verify_token)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return lead

@api_router.put("/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, lead: LeadCreate, user: dict = Depends(verify_token)):
    existing = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lead not found")
    if existing.get("is_converted"):
        raise HTTPException(status_code=400, detail="Cannot edit converted lead")
    
    # Calculate amounts
    products = []
    total_amount = 0
    for p in lead.products:
        amount = p.quantity * p.price
        products.append(ProductItem(
            product=p.product,
            part_number=p.part_number,
            category=p.category,
            quantity=p.quantity,
            price=p.price,
            amount=round(amount, 2)
        ))
        total_amount += amount
    
    update_data = lead.model_dump(exclude={"products", "tender_document", "working_sheet"})
    update_data["products"] = [p.model_dump() for p in products]
    update_data["total_amount"] = round(total_amount, 2)
    
    # Preserve existing document references - don't overwrite with None
    # Documents are managed via separate upload/delete endpoints
    
    await db.leads.update_one({"id": lead_id}, {"$set": update_data})
    updated = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return updated

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, user: dict = Depends(verify_token)):
    result = await db.leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"message": "Lead deleted"}

# ============== TENDER DOCUMENT UPLOAD ==============

@api_router.post("/leads/{lead_id}/upload-tender-document")
async def upload_tender_document(lead_id: str, file: UploadFile = File(...), user: dict = Depends(verify_token)):
    # Verify lead exists
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Validate file extension
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"File type not allowed. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    
    # Read and validate file size
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File size exceeds 25MB limit")
    
    # Generate unique filename
    unique_filename = f"{lead_id}_{uuid.uuid4().hex[:8]}{file_ext}"
    file_path = UPLOAD_DIR / unique_filename
    
    # Delete old tender document if exists
    if lead.get("tender_document"):
        old_file = UPLOAD_DIR / lead["tender_document"].split("/")[-1]
        if old_file.exists():
            old_file.unlink()
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Update lead with document path
    document_url = f"/api/uploads/{unique_filename}"
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"tender_document": document_url}}
    )
    
    return {"message": "Document uploaded successfully", "document_url": document_url, "filename": file.filename}

@api_router.delete("/leads/{lead_id}/tender-document")
async def delete_tender_document(lead_id: str, user: dict = Depends(verify_token)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if lead.get("tender_document"):
        old_file = UPLOAD_DIR / lead["tender_document"].split("/")[-1]
        if old_file.exists():
            old_file.unlink()
    
    await db.leads.update_one({"id": lead_id}, {"$set": {"tender_document": None}})
    return {"message": "Document deleted"}

# ============== WORKING SHEET UPLOAD ==============

@api_router.post("/leads/{lead_id}/upload-working-sheet")
async def upload_working_sheet(lead_id: str, file: UploadFile = File(...), user: dict = Depends(verify_token)):
    # Verify lead exists
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Validate file extension
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"File type not allowed. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    
    # Read and validate file size
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File size exceeds 25MB limit")
    
    # Generate unique filename with 'ws_' prefix to distinguish from tender docs
    unique_filename = f"ws_{lead_id}_{uuid.uuid4().hex[:8]}{file_ext}"
    file_path = UPLOAD_DIR / unique_filename
    
    # Delete old working sheet if exists
    if lead.get("working_sheet"):
        old_file = UPLOAD_DIR / lead["working_sheet"].split("/")[-1]
        if old_file.exists():
            old_file.unlink()
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Update lead with document path
    document_url = f"/api/uploads/{unique_filename}"
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"working_sheet": document_url}}
    )
    
    return {"message": "Working sheet uploaded successfully", "document_url": document_url, "filename": file.filename}

@api_router.delete("/leads/{lead_id}/working-sheet")
async def delete_working_sheet(lead_id: str, user: dict = Depends(verify_token)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if lead.get("working_sheet"):
        old_file = UPLOAD_DIR / lead["working_sheet"].split("/")[-1]
        if old_file.exists():
            old_file.unlink()
    
    await db.leads.update_one({"id": lead_id}, {"$set": {"working_sheet": None}})
    return {"message": "Working sheet deleted"}

@api_router.get("/uploads/{filename}")
async def serve_upload(filename: str):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path, filename=filename)

class ConvertLeadRequest(BaseModel):
    proforma_invoice_number: str

@api_router.post("/leads/{lead_id}/convert", response_model=ProformaInvoice)
async def convert_lead(lead_id: str, request: ConvertLeadRequest, user: dict = Depends(verify_token)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    if lead.get("is_converted"):
        raise HTTPException(status_code=400, detail="Lead already converted")
    
    proforma_number = request.proforma_invoice_number or lead.get("proforma_invoice_number")
    if not proforma_number:
        raise HTTPException(status_code=400, detail="Proforma Invoice Number is required for conversion")
    
    # Create Proforma Invoice with all lead data including documents
    proforma = ProformaInvoice(
        proforma_invoice_number=proforma_number,
        customer_id=lead["customer_id"],
        customer_name=lead["customer_name"],
        date=lead["date"],
        products=lead["products"],
        total_amount=lead["total_amount"],
        tender_document=lead.get("tender_document"),
        working_sheet=lead.get("working_sheet"),
        lead_id=lead_id
    )
    
    await db.proforma_invoices.insert_one(proforma.model_dump())
    
    # Mark lead as converted
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"is_converted": True, "proforma_invoice_number": proforma_number}}
    )
    
    return proforma

@api_router.post("/leads/bulk-upload")
async def bulk_upload_leads(file: UploadFile = File(...), user: dict = Depends(verify_token)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    leads_created = 0
    errors = []
    
    # Group rows by Customer Name + Proforma Invoice Number
    # New column order: Customer Name, PI No, Date, Product, Part Number, Category, Qty, Price, Follow-up, Remark
    lead_data = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        try:
            customer_name = str(row[0]).strip()
            proforma_number = str(row[1]).strip() if row[1] else ""
            
            # Create grouping key: prioritize PI number if exists, otherwise use customer name
            group_key = proforma_number if proforma_number else customer_name
            
            customer = await db.customers.find_one({"customer_name": {"$regex": f"^{customer_name}$", "$options": "i"}}, {"_id": 0})
            if not customer:
                errors.append(f"Row {row_idx}: Customer '{customer_name}' not found")
                continue
            
            if group_key not in lead_data:
                lead_data[group_key] = {
                    "customer_id": customer["id"],
                    "customer_name": customer["customer_name"],
                    "proforma_invoice_number": proforma_number if proforma_number else None,
                    "date": str(row[2]) if row[2] else datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "follow_up_date": str(row[8]) if len(row) > 8 and row[8] else None,
                    "remark": str(row[9]) if len(row) > 9 and row[9] else None,
                    "products": []
                }
            
            # Add product to the lead with part_number
            # Columns: Product(3), Part Number(4), Category(5), Quantity(6), Price(7)
            quantity = float(row[6]) if len(row) > 6 and row[6] else 0
            price = float(row[7]) if len(row) > 7 and row[7] else 0
            lead_data[group_key]["products"].append({
                "product": str(row[3]) if len(row) > 3 and row[3] else "",
                "part_number": str(row[4]) if len(row) > 4 and row[4] else None,
                "category": str(row[5]) if len(row) > 5 and row[5] else "",
                "quantity": quantity,
                "price": price,
                "amount": round(quantity * price, 2)
            })
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    # Create leads with grouped products
    for group_key, data in lead_data.items():
        try:
            total_amount = sum(p["amount"] for p in data["products"])
            lead = Lead(
                customer_id=data["customer_id"],
                customer_name=data["customer_name"],
                proforma_invoice_number=data["proforma_invoice_number"],
                date=data["date"],
                products=data["products"],
                follow_up_date=data["follow_up_date"],
                remark=data["remark"],
                total_amount=round(total_amount, 2)
            )
            await db.leads.insert_one(lead.model_dump())
            leads_created += 1
        except Exception as e:
            errors.append(f"Lead '{group_key}': {str(e)}")
    
    return {"created": leads_created, "errors": errors}

@api_router.get("/leads/template/download")
async def download_lead_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["Customer Name*", "Proforma Invoice No", "Date*", "Product*", "Part Number", "Category*", "Quantity*", "Price*", "Follow-up Date", "Remark"]
    ws.append(headers)
    # Example: Same PI number groups into one lead with 2 products
    ws.append(["John Doe", "PI-001", "2024-01-15", "Widget A", "PN-001", "Electronics", 10, 100, "2024-01-20", "Initial inquiry"])
    ws.append(["John Doe", "PI-001", "2024-01-15", "Widget B", "PN-002", "Electronics", 5, 200, "2024-01-20", "Initial inquiry"])
    # Example: Same customer without PI creates one lead with 1 product
    ws.append(["Jane Smith", "", "2024-01-16", "Gadget X", "PN-003", "Hardware", 3, 150, "2024-01-25", "Follow up needed"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lead_template.xlsx"}
    )

# ============== PROFORMA INVOICES ==============

@api_router.get("/proforma-invoices", response_model=List[ProformaInvoice])
async def get_proforma_invoices(
    customer_name: Optional[str] = None,
    category: Optional[str] = None,
    user: dict = Depends(verify_token)
):
    query = {}
    if customer_name:
        query["customer_name"] = {"$regex": customer_name, "$options": "i"}
    if category:
        query["products.category"] = {"$regex": category, "$options": "i"}
    
    invoices = await db.proforma_invoices.find(query, {"_id": 0}).to_list(1000)
    return invoices

@api_router.get("/proforma-invoices/{invoice_id}", response_model=ProformaInvoice)
async def get_proforma_invoice(invoice_id: str, user: dict = Depends(verify_token)):
    invoice = await db.proforma_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Proforma Invoice not found")
    
    # Real-time sync: Fetch latest data from linked lead
    if invoice.get("lead_id"):
        lead = await db.leads.find_one({"id": invoice["lead_id"]}, {"_id": 0})
        if lead:
            # Sync products, documents, and amounts from lead
            synced_data = {
                "products": lead.get("products", invoice.get("products")),
                "total_amount": lead.get("total_amount", invoice.get("total_amount")),
                "tender_document": lead.get("tender_document"),
                "working_sheet": lead.get("working_sheet"),
                "customer_name": lead.get("customer_name", invoice.get("customer_name"))
            }
            # Update invoice in database with synced data
            await db.proforma_invoices.update_one(
                {"id": invoice_id},
                {"$set": synced_data}
            )
            # Return updated invoice
            invoice = await db.proforma_invoices.find_one({"id": invoice_id}, {"_id": 0})
    
    return invoice

@api_router.put("/proforma-invoices/{invoice_id}", response_model=ProformaInvoice)
async def update_proforma_invoice(invoice_id: str, products: List[ProductItem], user: dict = Depends(verify_token)):
    existing = await db.proforma_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Proforma Invoice not found")
    
    total_amount = sum(p.quantity * p.price for p in products)
    updated_products = []
    for p in products:
        updated_products.append(ProductItem(
            product=p.product,
            part_number=p.part_number,
            category=p.category,
            quantity=p.quantity,
            price=p.price,
            amount=round(p.quantity * p.price, 2)
        ).model_dump())
    
    await db.proforma_invoices.update_one(
        {"id": invoice_id},
        {"$set": {"products": updated_products, "total_amount": round(total_amount, 2)}}
    )
    updated = await db.proforma_invoices.find_one({"id": invoice_id}, {"_id": 0})
    return updated

@api_router.delete("/proforma-invoices/{invoice_id}")
async def delete_proforma_invoice(invoice_id: str, user: dict = Depends(verify_token)):
    result = await db.proforma_invoices.delete_one({"id": invoice_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proforma Invoice not found")
    return {"message": "Proforma Invoice deleted"}

# ============== PURCHASE ORDERS ==============

@api_router.post("/purchase-orders", response_model=PurchaseOrder)
async def create_purchase_order(po: PurchaseOrderCreate, user: dict = Depends(verify_token)):
    # Validate proforma if linked
    if po.purpose == "linked" and po.proforma_invoice_id:
        proforma = await db.proforma_invoices.find_one({"id": po.proforma_invoice_id}, {"_id": 0})
        if not proforma:
            raise HTTPException(status_code=400, detail="Proforma Invoice not found")
    
    # Calculate amounts for each product and total
    products = []
    total_amount = 0
    for p in po.products:
        amount = p.quantity * p.price
        products.append(POProductItem(
            product=p.product,
            category=p.category,
            quantity=p.quantity,
            price=p.price,
            amount=round(amount, 2)
        ))
        total_amount += amount
    
    po_obj = PurchaseOrder(
        purchase_order_number=po.purchase_order_number,
        date=po.date,
        vendor_name=po.vendor_name,
        purpose=po.purpose,
        proforma_invoice_id=po.proforma_invoice_id,
        proforma_invoice_number=po.proforma_invoice_number,
        products=products,
        total_amount=round(total_amount, 2)
    )
    
    await db.purchase_orders.insert_one(po_obj.model_dump())
    return po_obj

@api_router.get("/purchase-orders", response_model=List[PurchaseOrder])
async def get_purchase_orders(
    vendor_name: Optional[str] = None,
    category: Optional[str] = None,
    date: Optional[str] = None,
    purpose: Optional[str] = None,
    user: dict = Depends(verify_token)
):
    query = {}
    if vendor_name:
        query["vendor_name"] = {"$regex": vendor_name, "$options": "i"}
    if category:
        query["$or"] = [
            {"products.category": {"$regex": category, "$options": "i"}},
            {"category": {"$regex": category, "$options": "i"}}
        ]
    if date:
        query["date"] = date
    if purpose:
        query["purpose"] = purpose
    
    orders = await db.purchase_orders.find(query, {"_id": 0}).to_list(1000)
    
    # Transform old format to new format for backward compatibility
    transformed_orders = []
    for order in orders:
        if "products" not in order or not order["products"]:
            # Old format - convert single product to products array
            if order.get("product"):
                order["products"] = [{
                    "product": order.get("product", ""),
                    "category": order.get("category", ""),
                    "quantity": order.get("quantity", 0),
                    "price": order.get("price", 0),
                    "amount": order.get("amount", 0)
                }]
                order["total_amount"] = order.get("amount", 0)
            else:
                order["products"] = []
                order["total_amount"] = 0
        transformed_orders.append(order)
    
    return transformed_orders

@api_router.get("/purchase-orders/{order_id}", response_model=PurchaseOrder)
async def get_purchase_order(order_id: str, user: dict = Depends(verify_token)):
    order = await db.purchase_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    # Transform old format to new format for backward compatibility
    if "products" not in order or not order["products"]:
        if order.get("product"):
            order["products"] = [{
                "product": order.get("product", ""),
                "category": order.get("category", ""),
                "quantity": order.get("quantity", 0),
                "price": order.get("price", 0),
                "amount": order.get("amount", 0)
            }]
            order["total_amount"] = order.get("amount", 0)
        else:
            order["products"] = []
            order["total_amount"] = 0
    
    return order

@api_router.put("/purchase-orders/{order_id}", response_model=PurchaseOrder)
async def update_purchase_order(order_id: str, po: PurchaseOrderCreate, user: dict = Depends(verify_token)):
    existing = await db.purchase_orders.find_one({"id": order_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    # Calculate amounts for each product and total
    products = []
    total_amount = 0
    for p in po.products:
        amount = p.quantity * p.price
        products.append(POProductItem(
            product=p.product,
            category=p.category,
            quantity=p.quantity,
            price=p.price,
            amount=round(amount, 2)
        ))
        total_amount += amount
    
    update_data = po.model_dump(exclude={"products"})
    update_data["products"] = [p.model_dump() for p in products]
    update_data["total_amount"] = round(total_amount, 2)
    
    await db.purchase_orders.update_one({"id": order_id}, {"$set": update_data})
    updated = await db.purchase_orders.find_one({"id": order_id}, {"_id": 0})
    return updated

@api_router.delete("/purchase-orders/{order_id}")
async def delete_purchase_order(order_id: str, user: dict = Depends(verify_token)):
    result = await db.purchase_orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    return {"message": "Purchase Order deleted"}

@api_router.post("/purchase-orders/bulk-upload")
async def bulk_upload_purchase_orders(file: UploadFile = File(...), user: dict = Depends(verify_token)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    orders_created = 0
    errors = []
    
    # Group rows by PO Number to support multiple products per PO
    po_data = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        try:
            po_number = str(row[0])
            if po_number not in po_data:
                proforma_id = None
                proforma_number = None
                purpose = str(row[3]).lower() if row[3] else "stock_in_sale"
                
                if purpose == "linked" and row[4]:
                    proforma = await db.proforma_invoices.find_one(
                        {"proforma_invoice_number": str(row[4])}, {"_id": 0}
                    )
                    if proforma:
                        proforma_id = proforma["id"]
                        proforma_number = proforma["proforma_invoice_number"]
                
                po_data[po_number] = {
                    "purchase_order_number": po_number,
                    "date": str(row[1]) if row[1] else datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "vendor_name": str(row[2]) if row[2] else "",
                    "purpose": purpose,
                    "proforma_invoice_id": proforma_id,
                    "proforma_invoice_number": proforma_number,
                    "products": []
                }
            
            # Add product to PO
            quantity = float(row[7]) if row[7] else 0
            price = float(row[8]) if row[8] else 0
            po_data[po_number]["products"].append({
                "product": str(row[5]) if row[5] else "",
                "category": str(row[6]) if row[6] else "",
                "quantity": quantity,
                "price": price,
                "amount": round(quantity * price, 2)
            })
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    # Create POs with their products
    for po_number, data in po_data.items():
        try:
            total_amount = sum(p["amount"] for p in data["products"])
            po = PurchaseOrder(
                **{k: v for k, v in data.items() if k != "products"},
                products=data["products"],
                total_amount=round(total_amount, 2)
            )
            await db.purchase_orders.insert_one(po.model_dump())
            orders_created += 1
        except Exception as e:
            errors.append(f"PO {po_number}: {str(e)}")
    
    return {"created": orders_created, "errors": errors}

@api_router.get("/purchase-orders/template/download")
async def download_po_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"
    headers = ["PO Number*", "Date*", "Vendor Name*", "Purpose (linked/stock_in_sale)*", "Proforma Invoice No", "Product*", "Category*", "Quantity*", "Price*"]
    ws.append(headers)
    # Example: PO-001 with 2 products, PO-002 with 1 product
    ws.append(["PO-001", "2024-01-15", "Vendor ABC", "linked", "PI-001", "Widget A", "Electronics", 5, 80])
    ws.append(["PO-001", "2024-01-15", "Vendor ABC", "linked", "PI-001", "Widget B", "Electronics", 3, 100])
    ws.append(["PO-002", "2024-01-16", "Vendor XYZ", "stock_in_sale", "", "Widget C", "Hardware", 10, 50])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchase_order_template.xlsx"}
    )

# ============== MARGIN CALCULATOR ==============

@api_router.get("/margin-calculator")
async def get_margin_data(user: dict = Depends(verify_token)):
    # Get proforma invoices with linked purchase orders
    proforma_list = await db.proforma_invoices.find({}, {"_id": 0}).to_list(1000)
    
    margin_data = []
    for pf in proforma_list:
        linked_orders = await db.purchase_orders.find(
            {"proforma_invoice_id": pf["id"]}, {"_id": 0}
        ).to_list(100)
        
        if not linked_orders:
            continue
        
        for po in linked_orders:
            # Use total_amount for POs with multiple products
            po_amount = po.get("total_amount", po.get("amount", 0))
            remaining = pf["total_amount"] - po_amount
            # Get saved freight from margin collection
            margin_doc = await db.margins.find_one(
                {"proforma_invoice_id": pf["id"], "purchase_order_id": po["id"]},
                {"_id": 0}
            )
            freight = margin_doc.get("freight_amount", 0) if margin_doc else 0
            margin_amount = remaining - freight
            
            margin_data.append({
                "proforma_invoice_number": pf["proforma_invoice_number"],
                "proforma_invoice_id": pf["id"],
                "proforma_total_amount": pf["total_amount"],
                "purchase_order_number": po["purchase_order_number"],
                "purchase_order_id": po["id"],
                "purchase_order_amount": po_amount,
                "remaining_amount": round(remaining, 2),
                "freight_amount": freight,
                "margin_amount": round(margin_amount, 2)
            })
    
    return margin_data

@api_router.put("/margin-calculator/{proforma_id}/{po_id}")
async def update_margin_freight(
    proforma_id: str,
    po_id: str,
    margin_update: MarginUpdate,
    user: dict = Depends(verify_token)
):
    # Upsert margin data
    await db.margins.update_one(
        {"proforma_invoice_id": proforma_id, "purchase_order_id": po_id},
        {"$set": {"freight_amount": margin_update.freight_amount}},
        upsert=True
    )
    return {"message": "Freight updated"}

# ============== GEM BID CRM MODULE ==============
# Completely separate module with its own authentication and data

# GEM BID Models
class GemBidStatusUpdate(BaseModel):
    status: str
    timestamp: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class GemBidBase(BaseModel):
    Firm_name: Optional[str] = None
    gem_bid_no: str
    Bid_details: Optional[str] = None
    description: Optional[str] = None
    start_date: str
    end_date: str
    emd_amount: float
    quantity: float
    city: Optional[str] = None
    department: Optional[str] = None
    item_category: Optional[str] = None
    epbg_percentage: Optional[float] = None
    epbg_month: Optional[int] = None
    status: str = "Shortlisted"

class GemBidCreate(GemBidBase):
    pass

class GemBid(GemBidBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status_history: List[GemBidStatusUpdate] = []
    documents: List[dict] = []  # List of {filename, url, uploaded_at}
    created_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# GEM BID Valid Statuses
GEM_BID_STATUSES = [
    "Shortlisted",
    "Participated",
    "Technical Evaluation",
    "RA",
    "Rejected",
    "Bid Awarded",
    "Supply Order Received",
    "Material Procurement",
    "Order Complete"
]

# GEM BID Order Model
class GemOrderItem(BaseModel):
    sku: str
    vendor: str
    price: float
    quantity: float
    invoice_value: float
    advance_paid: float
    remaining_amount: float = 0
    date: str
    delivery_date: str

class GemOrderBase(BaseModel):
    gem_bid_no: str
    items: List[GemOrderItem] = []

class GemOrderCreate(GemOrderBase):
    pass

class GemOrder(GemOrderBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# GEM BID Authentication
def verify_gem_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Verify JWT token for GEM BID CRM (separate from main CRM)"""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("system") != "gem_bid":
            raise HTTPException(status_code=401, detail="Invalid token for GEM BID CRM")
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

@api_router.post("/gem-bid/auth/login")
async def gem_bid_login(request: LoginRequest):
    # Hardcoded fallback from environment
    HARDCODED_GEM = {"email": GEM_BID_USER_EMAIL, "password": GEM_BID_USER_PASSWORD, "name": "Yash Bora"}
    
    try:
        user = await db.users.find_one({
            "email": request.email,
            "password": request.password,
            "system": "gem_bid"
        })
        if user:
            token_data = {
                "email": user["email"],
                "name": user["name"],
                "system": "gem_bid",
                "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
            }
            token = jwt.encode(token_data, SECRET_KEY, algorithm=ALGORITHM)
            return LoginResponse(token=token, user={"email": user["email"], "name": user["name"]})
    except Exception as e:
        logger.error(f"Database gem-bid login error: {e}")

    # Check hardcoded fallback
    if request.email == HARDCODED_GEM["email"] and request.password == HARDCODED_GEM["password"]:
        token_data = {
            "email": HARDCODED_GEM["email"],
            "name": HARDCODED_GEM["name"],
            "system": "gem_bid",
            "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        }
        token = jwt.encode(token_data, SECRET_KEY, algorithm=ALGORITHM)
        return LoginResponse(token=token, user={"email": HARDCODED_GEM["email"], "name": HARDCODED_GEM["name"]})
    
    raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.get("/gem-bid/scheduler/status")
async def get_scheduler_status_endpoint(user: dict = Depends(verify_gem_token)):
    """Get the status of the bid reminder scheduler"""
    status = get_scheduler_status()
    return status

# GEM BID CRUD Endpoints
@api_router.get("/gem-bid/bids", response_model=List[GemBid])
async def get_gem_bids(status_filter: Optional[str] = None, user: dict = Depends(verify_gem_token)):
    query = {}
    if status_filter:
        query["status"] = status_filter
    bids = await db.gem_bids.find(query, {"_id": 0}).sort("created_date", -1).to_list(1000)
    return bids

@api_router.get("/gem-bid/bids/new")
async def get_new_bids(user: dict = Depends(verify_gem_token)):
    """Get all bids except 'Bid Awarded', 'Supply Order Received', 'Material Procurement', 'Order Complete'"""
    bids = await db.gem_bids.find(
        {"status": {"$nin": ["Bid Awarded", "Supply Order Received", "Material Procurement", "Order Complete"]}},
        {"_id": 0}
    ).sort("created_date", -1).to_list(1000)
    return bids

@api_router.get("/gem-bid/bids/completed")
async def get_completed_bids(user: dict = Depends(verify_gem_token)):
    """Get bids with statuses: 'Bid Awarded', 'Supply Order Received', 'Material Procurement', 'Order Complete'"""
    bids = await db.gem_bids.find(
        {"status": {"$in": ["Bid Awarded", "Supply Order Received", "Material Procurement", "Order Complete"]}},
        {"_id": 0}
    ).sort("created_date", -1).to_list(1000)
    return bids

@api_router.get("/gem-bid/bids/{bid_id}", response_model=GemBid)
async def get_gem_bid(bid_id: str, user: dict = Depends(verify_gem_token)):
    bid = await db.gem_bids.find_one({"id": bid_id}, {"_id": 0})
    if not bid:
        raise HTTPException(status_code=404, detail="Bid not found")
    return bid

@api_router.post("/gem-bid/bids", response_model=GemBid)
async def create_gem_bid(bid: GemBidCreate, user: dict = Depends(verify_gem_token)):
    # Validate status
    if bid.status not in GEM_BID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(GEM_BID_STATUSES)}")
    
    bid_obj = GemBid(
        **bid.model_dump(),
        status_history=[GemBidStatusUpdate(status=bid.status)]
    )
    await db.gem_bids.insert_one(bid_obj.model_dump())
    return bid_obj

@api_router.put("/gem-bid/bids/{bid_id}", response_model=GemBid)
async def update_gem_bid(bid_id: str, bid: GemBidCreate, user: dict = Depends(verify_gem_token)):
    existing = await db.gem_bids.find_one({"id": bid_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Bid not found")
    
    # Validate status
    if bid.status not in GEM_BID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(GEM_BID_STATUSES)}")
    
    update_data = bid.model_dump()
    
    # If status changed, add to history
    if bid.status != existing.get("status"):
        status_history = existing.get("status_history", [])
        status_history.append(GemBidStatusUpdate(status=bid.status).model_dump())
        update_data["status_history"] = status_history
    
    await db.gem_bids.update_one({"id": bid_id}, {"$set": update_data})
    updated = await db.gem_bids.find_one({"id": bid_id}, {"_id": 0})
    return updated

@api_router.patch("/gem-bid/bids/{bid_id}/status")
async def update_gem_bid_status(bid_id: str, status: str, user: dict = Depends(verify_gem_token)):
    """Quick status update endpoint"""
    if status not in GEM_BID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(GEM_BID_STATUSES)}")
    
    existing = await db.gem_bids.find_one({"id": bid_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Bid not found")
    
    # Add to status history
    status_history = existing.get("status_history", [])
    status_history.append(GemBidStatusUpdate(status=status).model_dump())
    
    await db.gem_bids.update_one(
        {"id": bid_id},
        {"$set": {"status": status, "status_history": status_history}}
    )
    return {"message": f"Status updated to {status}"}

@api_router.delete("/gem-bid/bids/{bid_id}")
async def delete_gem_bid(bid_id: str, user: dict = Depends(verify_gem_token)):
    # Delete associated documents
    bid = await db.gem_bids.find_one({"id": bid_id}, {"_id": 0})
    if bid:
        for doc in bid.get("documents", []):
            file_path = GEM_BID_UPLOAD_DIR / doc["url"].split("/")[-1]
            if file_path.exists():
                file_path.unlink()
    
    result = await db.gem_bids.delete_one({"id": bid_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bid not found")
    return {"message": "Bid deleted"}

# GEM BID Document Upload
@api_router.post("/gem-bid/bids/{bid_id}/documents")
async def upload_gem_bid_document(bid_id: str, file: UploadFile = File(...), user: dict = Depends(verify_gem_token)):
    bid = await db.gem_bids.find_one({"id": bid_id}, {"_id": 0})
    if not bid:
        raise HTTPException(status_code=404, detail="Bid not found")
    
    # Validate file extension
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type not allowed. Allowed: {', '.join(ALLOWED_EXTENSIONS)}")
    
    # Read and validate file size
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File size exceeds 25MB limit")
    
    # Save file
    unique_filename = f"gem_{bid_id}_{uuid.uuid4().hex[:8]}{file_ext}"
    file_path = GEM_BID_UPLOAD_DIR / unique_filename
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Add document reference to bid
    document_url = f"/api/gem-bid/uploads/{unique_filename}"
    documents = bid.get("documents", [])
    documents.append({
        "filename": file.filename,
        "url": document_url,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    })
    
    await db.gem_bids.update_one({"id": bid_id}, {"$set": {"documents": documents}})
    return {"message": "Document uploaded", "document_url": document_url, "filename": file.filename}

@api_router.delete("/gem-bid/bids/{bid_id}/documents/{doc_index}")
async def delete_gem_bid_document(bid_id: str, doc_index: int, user: dict = Depends(verify_gem_token)):
    bid = await db.gem_bids.find_one({"id": bid_id}, {"_id": 0})
    if not bid:
        raise HTTPException(status_code=404, detail="Bid not found")
    
    documents = bid.get("documents", [])
    if doc_index < 0 or doc_index >= len(documents):
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Delete file
    doc = documents[doc_index]
    file_path = GEM_BID_UPLOAD_DIR / doc["url"].split("/")[-1]
    if file_path.exists():
        file_path.unlink()
    
    # Remove from list
    documents.pop(doc_index)
    await db.gem_bids.update_one({"id": bid_id}, {"$set": {"documents": documents}})
    return {"message": "Document deleted"}

@api_router.get("/gem-bid/uploads/{filename}")
async def serve_gem_upload(filename: str):
    file_path = GEM_BID_UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path, filename=filename)

# GEM BID Excel Template Download
@api_router.get("/gem-bid/template/download")
async def download_gem_bid_template(user: dict = Depends(verify_gem_token)):
    wb = Workbook()
    ws = wb.active
    ws.title = "GEM Bids"
    headers = [
        "Firm Name", "Gem Bid No*", "Bid Details", "Description", "Start Date* (YYYY-MM-DD)", "End Date* (YYYY-MM-DD)",
        "EMD Amount*", "Quantity*", "City", "Department", "Item Category",
        "EPBG Percentage", "EPBG Month", "Status*"
    ]
    ws.append(headers)
    # Example row
    ws.append([
        "ABC Corp", "GEM/2024/B/001", "Detailed specs for office equipment", "Supply of Office Equipment", "2024-01-15", "2024-02-15",
        50000, 100, "Delhi", "Ministry of Finance", "Electronics",
        5, 12, "Shortlisted"
    ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gem_bid_template.xlsx"}
    )

# GEM BID Bulk Upload
@api_router.post("/gem-bid/bulk-upload")
async def bulk_upload_gem_bids(file: UploadFile = File(...), user: dict = Depends(verify_gem_token)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    bids_created = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[1]: # Gem Bid No is now at index 1
            continue
        try:
            # Validate required fields
            firm_name = str(row[0]).strip() if row[0] else None
            gem_bid_no = str(row[1]).strip()
            bid_details = str(row[2]).strip() if row[2] else None
            description = str(row[3]).strip() if row[3] else None
            start_date = str(row[4]).strip() if row[4] else None
            end_date = str(row[5]).strip() if row[5] else None
            emd_amount = float(row[6]) if row[6] else None
            quantity = float(row[7]) if row[7] else None
            
            # Map remaining fields
            city = str(row[8]).strip() if len(row) > 8 and row[8] else None
            department = str(row[9]).strip() if len(row) > 9 and row[9] else None
            item_category = str(row[10]).strip() if len(row) > 10 and row[10] else None
            epbg_percentage = float(row[11]) if len(row) > 11 and row[11] else None
            epbg_month = int(row[12]) if len(row) > 12 and row[12] else None
            status = str(row[13]).strip() if len(row) > 13 and row[13] else "Shortlisted"
            
            if not all([start_date, end_date, emd_amount is not None, quantity is not None]):
                errors.append(f"Row {row_idx}: Missing required fields")
                continue
            
            if status not in GEM_BID_STATUSES:
                status = "Shortlisted"
            
            bid = GemBid(
                Firm_name=firm_name,
                gem_bid_no=gem_bid_no,
                Bid_details=bid_details,
                description=description,
                start_date=start_date,
                end_date=end_date,
                emd_amount=emd_amount,
                quantity=quantity,
                city=city,
                department=department,
                item_category=item_category,
                epbg_percentage=epbg_percentage,
                epbg_month=epbg_month,
                status=status,
                status_history=[GemBidStatusUpdate(status=status)]
            )
            await db.gem_bids.insert_one(bid.model_dump())
            bids_created += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return {"created": bids_created, "errors": errors}

# GEM BID Statuses List (for dropdown)
@api_router.get("/gem-bid/statuses")
async def get_gem_bid_statuses(user: dict = Depends(verify_gem_token)):
    return GEM_BID_STATUSES

# GEM BID Orders Endpoints
@api_router.get("/gem-bid/orders", response_model=List[GemOrder])
async def get_gem_orders(user: dict = Depends(verify_gem_token)):
    orders = await db.gem_orders.find({}, {"_id": 0}).sort("created_date", -1).to_list(1000)
    
    # Handle backward compatibility for old single-SKU orders
    for order in orders:
        if "items" not in order or not order["items"]:
            # Synthesize an items list from root fields
            order["items"] = [{
                "sku": order.get("sku", "-"),
                "vendor": order.get("vendor", "-"),
                "price": order.get("price", 0),
                "quantity": order.get("quantity", 0),
                "invoice_value": order.get("invoice_value", 0),
                "advance_paid": order.get("advance_paid", 0),
                "remaining_amount": order.get("remaining_amount", 0),
                "date": order.get("date", ""),
                "delivery_date": order.get("delivery_date", "")
            }]
    return orders

@api_router.get("/gem-bid/orders/{order_id}", response_model=GemOrder)
async def get_gem_order(order_id: str, user: dict = Depends(verify_gem_token)):
    order = await db.gem_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
        
    # Handle backward compatibility
    if "items" not in order or not order["items"]:
        order["items"] = [{
            "sku": order.get("sku", "-"),
            "vendor": order.get("vendor", "-"),
            "price": order.get("price", 0),
            "quantity": order.get("quantity", 0),
            "invoice_value": order.get("invoice_value", 0),
            "advance_paid": order.get("advance_paid", 0),
            "remaining_amount": order.get("remaining_amount", 0),
            "date": order.get("date", ""),
            "delivery_date": order.get("delivery_date", "")
        }]
    return order

@api_router.post("/gem-bid/orders", response_model=GemOrder)
async def create_gem_order(order: GemOrderCreate, user: dict = Depends(verify_gem_token)):
    # Calculate remaining amount for each item
    for item in order.items:
        item.remaining_amount = round(item.invoice_value - item.advance_paid, 2)
        
    order_obj = GemOrder(
        **order.model_dump()
    )
    await db.gem_orders.insert_one(order_obj.model_dump())
    return order_obj

@api_router.put("/gem-bid/orders/{order_id}", response_model=GemOrder)
async def update_gem_order(order_id: str, order: GemOrderCreate, user: dict = Depends(verify_gem_token)):
    existing = await db.gem_orders.find_one({"id": order_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Calculate remaining amount for each item
    for item in order.items:
        item.remaining_amount = round(item.invoice_value - item.advance_paid, 2)
    
    update_data = order.model_dump()
    await db.gem_orders.update_one({"id": order_id}, {"$set": update_data})
    updated = await db.gem_orders.find_one({"id": order_id}, {"_id": 0})
    return updated

@api_router.delete("/gem-bid/orders/{order_id}")
async def delete_gem_order(order_id: str, user: dict = Depends(verify_gem_token)):
    result = await db.gem_orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"message": "Order deleted"}

# ============== ROOT ==============

@api_router.get("/")
async def root():
    return {"message": "CRM API Running"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    shutdown_scheduler()
    mongo_client.close()


